import webbrowser,sys
import pyperclip
import requests
import re
import os
import openpyxl  
import xlwt

def openBrowser():
    
    webbrowser.open('https://www.google.com.mx')
    
def commandLineArgs():
    if len(sys.argv) > 1:
        address = ' '.join(sys.argv[1:])
    else:
        address.pyperclip.paste()
            
        

def firstRequest():
    res = requests.get('https://www.meteored.mx/clima_Durango-America\+Norte-Mexico-Durango-MMDO-1-22326.html')
    texto = res.text
    pattern = re.compile('[Dd]urango[a-z0-9]+')
    texto_found = pattern.findall(texto)
    return texto_found
print(firstRequest())
       
       
def readExcelFile():
    os.chdir('C:\\Users\\uziel.buendia\\Downloads\\')
    workbook1 = openpyxl.load_workbook('BVTfile.xlsx')
    w1=workbook1.active
    for row in range(0,w1.max_row):
        for colc in w1.iter_cols(1,w1.max_column):
            if colc[row].value =='Versioning':
                
                updateValue(colc[row].value)
            print(colc[row].value)

    
def updateValue(data_found):
    book1 = xlwt.Workbook(encoding='utf-8')
    sheet1 = book1.add_sheet('MyFirstSheetUSB')
    sheet1.write(0,0,data_found)
    book1.save('excelwrite1.xslx')
    print("test")
            
       